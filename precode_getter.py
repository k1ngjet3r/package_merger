from openpyxl import load_workbook, Workbook

def dict_to_str(d):
    return str(d).replace("'", '"')


def precode_getter(wb, sheet):
    ws = load_workbook(wb)[sheet]

    pkg_precode_dict = {}

    for i in ws.iter_rows(max_col=2, values_only=True):
        pkg_precode_dict[i[0]] = i[1].split('\n\n')[-2].split('\n')

    print(dict_to_str(pkg_precode_dict))

# s = 'Test environment:\nBench & C-ATS 7.4 & MY22GB10TB HMI (You can get C-ATS 7.4 trail version in FTP: ftp://aliceliu@172.16.40.64/C-ATS%20trial/Installation/R7.1.0)\n1. Andriod phone model: Nexus 5 or Pixel 3 (Android 10)\nGo to Settings->WIFI Hotspot on CSM, change WIFI name to "myGMC" (Make sure the wifi name without space)\n2. Set BT time in phone is same as CSM time\n\nea-bt1c1-unpair\nea-bt-nobt\nea-rmtp4-normal\nea-usb-nousb\nea-sd-nosd\nea-aap-noaaphone\nea-carp-noiphone\nea-bt2c2-unpair\nsc-wifi-register\nea-wifi-antenna24g\nac-gas-testfang\npr-gas-toronto\npr-ecud-onstar\n\nThe precode definition refer to CNT-CATS-PLAN-Precondition_definition.xls'
# print(s.split('\n\n')[1].split('\n'))

def precode_conflict(wb, sheet):
    ws = load_workbook(wb)[sheet]

    conflict_dict =  {i[0]: i[1].split(',') for i in ws.iter_rows(max_col=2, values_only=True)}

    print(dict_to_str(conflict_dict))


if __name__ == '__main__':
    # precode_getter('pkg_precode.xlsx','13_inch')
    precode_conflict('C-ATS_LabDesign_Guideline.xlsx', 'Conflict')
