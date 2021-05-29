from openpyxl import load_workbook, Workbook
import json

def check_intersection(list1, list2):
    list1_as_set = set(list1)
    intersection = list1_as_set.intersection(list2)
    return len(intersection)

def open_json(filename):
    # For Windows
    with open('json\\' + filename) as f:
    # For MacOS
    # with open('json/' + filename) as f:
        return json.load(f)

conflict_dict = open_json('pkg_conflict.json')
pkg_def_dict = open_json('pkg_def.json')

class Pkg_merger:
    def __init__(self, tcid_wb):
        self.tcid_wb = load_workbook('xlsx/' + tcid_wb).active

    def script_with_precode(self):
        # ws_10 = load_workbook('xlsx\\Precode_10.xlsx')['BJ']
        ws_13 = load_workbook('xlsx\\Precode_13.xlsx')['BJ']

        tcid_ws = self.tcid_wb

        # for tcid in tcid_ws.iter_rows(max_col = 1, values_only=True):
        #     if tcid_list[0] != 'Original GM TC ID':
        
        tcid_list = [tcid[0].lower() for tcid in tcid_ws.iter_rows(max_col=1, values_only=True) if tcid[0] != 'Original GM TC ID' and tcid[0] != None]
        
        return {case[0].lower(): case[-1].split(',') for case in ws_13.iter_rows(max_col=3, values_only=True) if case[0] in tcid_list}

    def conflicted_precode_dict(self):
        script_and_precode = self.script_with_precode()
        script_with_conflict_precode = {}
        
        for script in script_and_precode:
            conflict_list = []
            for precode in script_and_precode[script]:
                if precode in conflict_dict:
                    for j in conflict_dict[precode]:
                        conflict_list.append(j)
            script_with_conflict_precode[script] = conflict_list
        return script_with_conflict_precode

    def iter_pkg(self):
        script_and_precode = self.script_with_precode()
        scripts_conflict = open_json('case_with_conflict.json')
        pkg_precode = open_json('current_pkg.json')

        # Initialized the output file 
        wb = Workbook()
        ws = wb.active
        # adding header to the output file
        ws.append([pkg for pkg in pkg_precode])

        for tcid in script_and_precode:
            append_list = [tcid]
            if tcid in scripts_conflict:
                conflict_precode = scripts_conflict[tcid]
                for pkg in pkg_precode:
                    if check_intersection(conflict_precode, pkg_precode[pkg]) != 0:
                        append_list.append('X')
                    else:
                        append_list.append(check_intersection(script_and_precode[tcid], pkg_precode[pkg]))
            
            else:
                for pkg in pkg_precode:
                    append_list.append(check_intersection(script_and_precode[tcid], pkg_precode[pkg]))

            print(append_list)
            ws.append(append_list)
        
        wb.save('matching_result.xlsx')


if __name__ == '__main__':
    script_ConflictPrecode = Pkg_merger('script_without_home.xlsx').iter_pkg()

    # print(len(script_ConflictPrecode))

    # output_wb = Workbook()
    # output_ws = output_wb.active
    
    # for tcid in script_ConflictPrecode:
    #     l = [tcid] + script_ConflictPrecode[tcid]
    #     output_ws.append(l)
    # output_wb.save('49_cases_with_precode.xlsx')